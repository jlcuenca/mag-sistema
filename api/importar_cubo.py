"""
Importador de Reporte Cubo 2025 — Carga datos reales a MAG Sistema
Lee las 3 hojas del Excel (RESUMEN, GENERAL, DETALLE) y puebla la BD.

Uso:
    python api/importar_cubo.py
"""
import os
import sys
import openpyxl
from datetime import datetime

# Setup path
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, BASE_DIR)

from api.database import (
    init_db, SessionLocal, engine, Base,
    Agente, Producto, Poliza, Recibo, Segmento,
    Meta, Importacion
)
from api.rules import (
    normalizar_poliza, calcular_mystatus, agrupar_segmento,
    clasificar_cy, es_reexpedicion, extraer_raiz_poliza
)

# ── Configuración ──────────────────────────────────────────────────
CUBO_PATH = os.path.join(BASE_DIR, "fuentes", "Reporte_Cubo_2025_ALL (3).xlsx")

# Segmentos a crear
SEGMENTOS_CATALOGO = [
    ("ALFA TOP INTEGRAL",  "ALFA", 1),
    ("ALFA TOP COMBINADO", "ALFA", 2),
    ("ALFA TOP",           "ALFA", 3),
    ("ALFA INTEGRAL",      "ALFA", 4),
    ("ALFA/BETA",          "ALFA", 5),
    ("BETA1",              "BETA", 10),
    ("BETA2",              "BETA", 11),
    ("OMEGA",              "OMEGA", 20),
]


def safe_str(v) -> str:
    """Convierte un valor a string, retorna '' si None."""
    if v is None:
        return ""
    return str(v).strip()


def safe_float(v) -> float:
    """Convierte a float, retorna 0.0 si falla."""
    if v is None:
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def safe_int(v) -> int:
    """Convierte a int, retorna 0 si falla."""
    if v is None:
        return 0
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return 0


def parse_date(v) -> str:
    """Convierte un valor de fecha a string YYYY-MM-DD."""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s or s == "None":
        return ""
    # Ya en formato ISO
    if len(s) >= 10 and s[4] == '-':
        return s[:10]
    return s[:10]


def read_header_row(ws, header_row=5):
    """Lee la fila de encabezados y retorna un dict col_idx -> nombre."""
    headers = {}
    for col_idx, cell in enumerate(ws[header_row], 1):
        if cell.value:
            headers[col_idx] = str(cell.value).strip()
    return headers


def row_to_dict(row, headers):
    """Convierte una fila a dict usando los headers."""
    d = {}
    for col_idx, cell in enumerate(row, 1):
        key = headers.get(col_idx)
        if key:
            d[key] = cell.value
    return d


def main():
    print("=" * 70)
    print("  IMPORTADOR REPORTE CUBO 2025 — MAG Sistema")
    print("=" * 70)

    if not os.path.exists(CUBO_PATH):
        print(f"❌ No se encontró: {CUBO_PATH}")
        return

    print(f"📂 Archivo: {CUBO_PATH}")
    print(f"📏 Tamaño: {os.path.getsize(CUBO_PATH):,} bytes")

    # ── Borrar BD vieja y recrear ──────────────────────────────────
    from api.database import DB_PATH
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
        print(f"🗑️  BD anterior eliminada: {DB_PATH}")

    init_db()
    db = SessionLocal()

    try:
        # ── 1. Crear catálogo de segmentos ─────────────────────────
        print("\n📋 Creando catálogo de segmentos...")
        segmentos_map = {}
        for nombre, agrupado, orden in SEGMENTOS_CATALOGO:
            seg = Segmento(nombre=nombre, agrupado=agrupado, orden=orden)
            db.add(seg)
            db.flush()
            segmentos_map[nombre] = seg.id
        print(f"   → {len(segmentos_map)} segmentos creados")

        # ── 2. Crear producto base ──────────────────────────────────
        print("\n📦 Creando catálogo de productos...")
        # Solo GMM 34 FLEX PLUS aparece en el cubo
        prod_gmm = Producto(ramo_codigo=34, ramo_nombre="GASTOS MEDICOS MAYORES INDIVIDUAL",
                           plan="FLEX PLUS", gama=None)
        db.add(prod_gmm)
        prod_vida = Producto(ramo_codigo=11, ramo_nombre="VIDA",
                            plan="VIDA Y AHORRO", gama=None)
        db.add(prod_vida)
        db.flush()
        print(f"   → Productos: GMM (id={prod_gmm.id}), VIDA (id={prod_vida.id})")

        # ── 3. Leer Excel ──────────────────────────────────────────
        print(f"\n📖 Leyendo Excel...")
        wb = openpyxl.load_workbook(CUBO_PATH, data_only=True)
        print(f"   Hojas: {wb.sheetnames}")

        # ── 4. Procesar hoja RESUMEN → Agentes ─────────────────────
        print("\n👥 Procesando hoja RESUMEN → Agentes...")
        ws_resumen = wb["RESUMEN"]

        # Headers en fila 5
        headers_res = read_header_row(ws_resumen, 5)
        print(f"   Headers: {list(headers_res.values())}")

        agentes_map = {}  # codigo -> agente_id
        agentes_count = 0
        for row in ws_resumen.iter_rows(min_row=6, values_only=False):
            d = row_to_dict(row, headers_res)
            codigo = safe_str(d.get("CLAVE AGENTE"))
            nombre = safe_str(d.get("NOMBRE COMPLETO"))

            if not codigo or not nombre:
                continue

            segmento_nombre = safe_str(d.get("SEGMENTO"))
            gestion = safe_str(d.get("GESTION COMERCIAL"))
            lider = safe_str(d.get("LIDER"))

            seg_id = segmentos_map.get(segmento_nombre)
            seg_agrupado = agrupar_segmento(segmento_nombre)

            agente = Agente(
                codigo_agente=codigo,
                nombre_completo=nombre,
                segmento_id=seg_id,
                segmento_nombre=segmento_nombre,
                segmento_agrupado=seg_agrupado,
                gestion_comercial=gestion,
                lider_codigo=lider,
                nombre_promotoria="MAG",
                situacion="ACTIVO",
            )
            db.add(agente)
            db.flush()
            agentes_map[codigo] = agente.id
            agentes_count += 1

        print(f"   → {agentes_count} agentes creados")

        # ── 5. Procesar hoja GENERAL → Pólizas ─────────────────────
        print("\n📋 Procesando hoja GENERAL → Pólizas...")
        ws_general = wb["GENERAL"]
        headers_gen = read_header_row(ws_general, 5)
        print(f"   Headers: {list(headers_gen.values())[:10]}...")

        polizas_map = {}  # poliza_estandar -> poliza_id
        polizas_count = 0
        errors = []

        for row_idx, row in enumerate(ws_general.iter_rows(min_row=6, values_only=False), 6):
            d = row_to_dict(row, headers_gen)
            poliza_num = safe_str(d.get("POLIZA"))
            if not poliza_num:
                continue

            agente_codigo = safe_str(d.get("AGENT") or d.get("AGENTE"))
            ramo = safe_int(d.get("RAMO"))
            plan = safe_str(d.get("PLAN"))
            contratante = safe_str(d.get("CONTRATANTE"))
            segmento = safe_str(d.get("SEGMENTO"))
            clasificacion = safe_str(d.get("CLASIFICACION"))
            forma_pago = safe_str(d.get("FORMA PAGO"))
            tipo_pago = safe_str(d.get("TIPO_PAGO"))
            nueva_flag = safe_int(d.get("NUEVA_POLIZA"))
            estatus = safe_str(d.get("ESTATUS"))
            detalle_estatus = safe_str(d.get("DETALLE_ESTATUS"))
            fecha_inicio = parse_date(d.get("INICIO VIGENCIA"))
            fecha_fin = parse_date(d.get("FIN VIGENCIA"))
            fecha_primer_pago = parse_date(d.get("FECHA_PRIMER_PAGO"))
            fecha_ultimo_pago = parse_date(d.get("FECHA_ULTIMO_PAGO"))
            anio_conf = safe_int(d.get("AÑO_CONF"))
            mes_conf_val = d.get("MES CONF")

            neta_total_contrato = safe_float(d.get("NETA_TOTAL_CONTRATO"))
            neta_acumulada = safe_float(d.get("NETA_TOTAL_ACUMULADO") or d.get("NETA_TOTAL_ACUMULADA"))
            neta_forma_pago = safe_float(d.get("NETA_SEGUN_FORMA_PAGO"))
            neta_sin_forma = safe_float(d.get("NETA_SIN_FORMA_PAGO"))
            neta_cancelacion = safe_float(d.get("NETA_CON_CANCELACION"))
            num_asegurados = safe_int(d.get("ASEGURADOS"))

            poliza_estandar = normalizar_poliza(poliza_num)
            agente_id = agentes_map.get(agente_codigo)
            producto_id = prod_gmm.id if ramo == 34 else prod_vida.id if ramo == 11 else None

            # Si el agente no existía en RESUMEN, crearlo
            if not agente_id and agente_codigo:
                nombre_agente = safe_str(d.get("NOMBRE_AGENTE"))
                ag = Agente(
                    codigo_agente=agente_codigo,
                    nombre_completo=nombre_agente or f"AGENTE {agente_codigo}",
                    segmento_nombre=segmento,
                    segmento_agrupado=agrupar_segmento(segmento),
                    nombre_promotoria="MAG",
                    situacion="ACTIVO",
                )
                db.add(ag)
                db.flush()
                agentes_map[agente_codigo] = ag.id
                agente_id = ag.id

            # Clasificar con reglas CY del cubo
            cy = clasificar_cy(clasificacion)

            # Calcular mystatus con los datos del cubo
            mystatus = calcular_mystatus(None, estatus, detalle_estatus)

            # Status simplificado para filtros
            if "PAGADA" in estatus.upper() or "AL CORRIENTE" in estatus.upper():
                status_recibo = "PAGADA"
            elif "CANCELADA" in estatus.upper():
                status_recibo = "CANCELADA"
            elif "ATRASADA" in estatus.upper():
                status_recibo = "ATRASADA"
            else:
                status_recibo = estatus

            # Determinar año y periodo de aplicación desde fecha_inicio
            anio_apli = int(fecha_inicio[:4]) if fecha_inicio and len(fecha_inicio) >= 4 else 2025
            periodo = fecha_inicio[:7] if fecha_inicio and len(fecha_inicio) >= 7 else f"{anio_apli}-01"

            # Reclasificar tipo_poliza basado en año real
            if anio_apli == 2025:
                if status_recibo in ("PAGADA", "AL CORRIENTE"):
                    tipo_final = "NUEVA"
                    es_nueva_final = True
                else:
                    tipo_final = "NUEVA"  # Canceladas del 2025 aún son NUEVA
                    es_nueva_final = False
            elif anio_apli == 2024:
                tipo_final = "SUBSECUENTE"
                es_nueva_final = False
            else:
                tipo_final = cy["tipo_poliza"]
                es_nueva_final = cy["es_nueva"]

            try:
                poliza = Poliza(
                    poliza_original=poliza_num,
                    poliza_estandar=poliza_estandar,
                    agente_id=agente_id,
                    producto_id=producto_id,
                    contratante_nombre=contratante,
                    num_asegurados=num_asegurados if num_asegurados > 0 else 1,
                    fecha_inicio=fecha_inicio or "2025-01-01",
                    fecha_fin=fecha_fin,
                    prima_neta=neta_total_contrato,
                    prima_total=neta_total_contrato,
                    forma_pago=forma_pago,
                    tipo_pago=tipo_pago,
                    status_recibo=status_recibo,
                    gama=None,  # En el cubo no viene la gama individual

                    # Clasificación
                    es_nueva=es_nueva_final,
                    tipo_poliza=tipo_final,
                    es_reexpedicion=es_reexpedicion(poliza_num),
                    mystatus=mystatus,

                    # Período
                    periodo_aplicacion=periodo,
                    anio_aplicacion=anio_apli,

                    # Campos nuevos del cubo
                    segmento=segmento,
                    gestion_comercial=safe_str(d.get("GESTION_COMERCIAL") or ""),
                    lider_codigo=safe_str(d.get("LIDER") or ""),
                    clasificacion_cy=clasificacion,
                    estatus_cubo=estatus,
                    estatus_detalle=detalle_estatus,
                    nueva_poliza_flag=nueva_flag,

                    # Multi-prima
                    neta_total_contrato=neta_total_contrato,
                    neta_acumulada=neta_acumulada,
                    neta_forma_pago=neta_forma_pago,
                    neta_sin_forma=neta_sin_forma,
                    neta_cancelacion=neta_cancelacion,

                    # Fechas cobranza
                    fecha_primer_pago=fecha_primer_pago,
                    fecha_ultimo_pago=fecha_ultimo_pago,

                    # Confirmación
                    anio_conf=anio_conf,
                    mes_conf=safe_int(mes_conf_val) if mes_conf_val else None,

                    fuente="REPORTE_CUBO",
                )
                db.add(poliza)
                db.flush()
                polizas_map[poliza_estandar] = poliza.id
                polizas_count += 1
            except Exception as e:
                errors.append(f"Fila {row_idx}: {poliza_num} - {e}")

        print(f"   → {polizas_count} pólizas creadas")
        if errors:
            print(f"   ⚠️ {len(errors)} errores:")
            for err in errors[:5]:
                print(f"      {err}")

        # ── 6. Procesar hoja DETALLE → Recibos ────────────────────
        print("\n💰 Procesando hoja DETALLE → Recibos...")
        ws_detalle = wb["DETALLE"]
        headers_det = read_header_row(ws_detalle, 5)
        print(f"   Headers: {list(headers_det.values())[:10]}...")

        recibos_count = 0
        for row in ws_detalle.iter_rows(min_row=6, values_only=False):
            d = row_to_dict(row, headers_det)
            poliza_num = safe_str(d.get("POLIZA"))
            if not poliza_num:
                continue

            poliza_estandar = normalizar_poliza(poliza_num)
            poliza_id = polizas_map.get(poliza_estandar)
            agente_codigo = safe_str(d.get("AGENTE") or d.get("AGENT"))

            recibo = Recibo(
                poliza_id=poliza_id,
                poliza_numero=poliza_num,
                fecha_recibo=parse_date(d.get("FECHA_RECIBO")),
                anio_apli=safe_int(d.get("AÑO_APLI")),
                mes_conf=safe_int(d.get("MES_CONF")),
                anio_conf=safe_int(d.get("AÑO_CONF")),
                comprobante=safe_str(d.get("COMPROBANTE")),
                neta_acumulada=safe_float(d.get("NETA_TOTAL_ACUMULADA") or d.get("NETA_TOTAL_ACUMULADO")),
                neta_forma_pago=safe_float(d.get("NETA_SEGUN_FORMA_PAGO")),
                neta_sin_forma=safe_float(d.get("NETA_SIN_FORMA_PAGO")),
                neta_cancelacion=safe_float(d.get("NETA_CON_CANCELACION")),
                agente_codigo=agente_codigo,
                nombre_agente=safe_str(d.get("NOMBRE_AGENTE")),
                ramo=safe_int(d.get("RAMO")),
                plan=safe_str(d.get("PLAN")),
                segmento=safe_str(d.get("SEGMENTO")),
                contratante=safe_str(d.get("CONTRATANTE")),
                estatus=safe_str(d.get("ESTATUS")),
                estatus_detalle=safe_str(d.get("DETALLE_ESTATUS")),
            )
            db.add(recibo)
            recibos_count += 1

            if recibos_count % 500 == 0:
                db.flush()

        print(f"   → {recibos_count} recibos creados")

        # ── 7. Crear metas de ejemplo ──────────────────────────────
        print("\n🎯 Creando metas 2025...")
        db.add(Meta(
            anio=2025, periodo=None,
            meta_polizas_vida=260, meta_prima_vida=16_841_996,
            meta_polizas_gmm=672, meta_asegurados_gmm=1019,
            meta_prima_gmm=15_054_843,
        ))

        # ── 8. Log de importación ──────────────────────────────────
        db.add(Importacion(
            tipo="REPORTE_CUBO",
            archivo_nombre=os.path.basename(CUBO_PATH),
            registros_procesados=polizas_count + recibos_count,
            registros_nuevos=polizas_count,
            registros_actualizados=0,
            registros_error=len(errors),
            errores_detalle="\n".join(errors[:20]) if errors else None,
            usuario="sistema",
        ))

        # ── Commit final ───────────────────────────────────────────
        db.commit()
        wb.close()

        # ── Resumen ────────────────────────────────────────────────
        print("\n" + "=" * 70)
        print("  ✅ IMPORTACIÓN COMPLETADA")
        print("=" * 70)
        print(f"  📊 Segmentos:  {len(segmentos_map)}")
        print(f"  👥 Agentes:    {agentes_count}")
        print(f"  📋 Pólizas:    {polizas_count}")
        print(f"  💰 Recibos:    {recibos_count}")
        print(f"  📦 Productos:  2")
        print(f"  ⚠️  Errores:    {len(errors)}")

        # Stats rápidas
        total_prima = db.query(Poliza).count()
        from sqlalchemy import func as sqf
        prima_total = db.query(sqf.sum(Poliza.neta_total_contrato)).scalar() or 0
        print(f"\n  💲 Prima neta total: ${prima_total:,.2f}")
        print(f"  📈 Total pólizas en BD: {total_prima}")

    except Exception as e:
        db.rollback()
        print(f"\n❌ Error fatal: {e}")
        import traceback
        traceback.print_exc()
    finally:
        db.close()


if __name__ == "__main__":
    main()
