
import openpyxl
import os

FILE_PATH = r"c:\Users\jlcue\Documents\mag-sistema\ref\Reporte 08 FEBRERO 2026 BASE VIDA Y GMM VF2 OK ULTIMO (1).xlsm"
TARGET_SUM = 2455490.3800000004

def find_target_sum():
    wb = openpyxl.load_workbook(FILE_PATH, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        if sheet_name == "RESUMEN VIDA Y GMM": continue
        ws = wb[sheet_name]
        print(f"Checking {sheet_name}...")
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=100, values_only=True)):
             for c_idx, val in enumerate(row):
                 if isinstance(val, (int, float)) and abs(val - TARGET_SUM) < 1.0:
                     print(f"FOUND MATCH IN {sheet_name}! Row {r_idx+1} Col {c_idx+1}")
    wb.close()

find_target_sum()
