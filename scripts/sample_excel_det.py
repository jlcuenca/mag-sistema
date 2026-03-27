
import openpyxl
import os

FILE_PATH = r"c:\Users\jlcue\Documents\mag-sistema\ref\Reporte 08 FEBRERO 2026 BASE VIDA Y GMM VF2 OK ULTIMO (1).xlsm"
SHEET_NAME = "DET 2025"

wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
ws = wb[SHEET_NAME]

print("POLIZA | AGENTE | MON | NETO | RAMO | MES")
for r in range(2, 20):
    p = ws.cell(row=r, column=1).value
    ag = ws.cell(row=r, column=2).value # 47968
    mon = ws.cell(row=r, column=12).value
    neto = ws.cell(row=r, column=11).value
    ramo = ws.cell(row=r, column=10).value
    mes = ws.cell(row=r, column=13).value
    print(f"{p} | {ag} | {mon} | {neto} | {ramo} | {mes}")
wb.close()
