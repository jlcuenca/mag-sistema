
import openpyxl
import os

FILE_PATH = r"c:\Users\jlcue\Documents\mag-sistema\ref\Reporte 08 FEBRERO 2026 BASE VIDA Y GMM VF2 OK ULTIMO (1).xlsm"

def get_agents():
    wb = openpyxl.load_workbook(FILE_PATH, read_only=True, data_only=True)
    ws = wb["RESUMEN VIDA Y GMM"]
    agents = []
    for r in range(10, 200):
        code = ws.cell(row=r, column=1).value
        if not code: break
        agents.append(str(code))
    wb.close()
    return agents

def find_new_match():
    agents = get_agents()
    wb = openpyxl.load_workbook(FILE_PATH, read_only=True, data_only=True)
    ws = wb["PRIMA PAGADA"]
    sum_new = 0
    sum_sub = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if i > 255000: break
        
        agente = str(row[0])
        anio = str(row[18])
        neta = float(row[10] or 0)
        is_new = row[34] # Col 35
        
        if agente in agents and anio == "2024" and str(row[2]).upper() == 'GAMMA FLEX IND.':
            if is_new == 1:
                sum_new += neta
            else:
                sum_sub += neta
                
    print(f"MAG 2024: New={sum_new:,.2f}, Sub={sum_sub:,.2f}")
    wb.close()

find_new_match()
