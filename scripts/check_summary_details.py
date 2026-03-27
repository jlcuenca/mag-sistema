
import openpyxl
import os

FILE_PATH = r"c:\Users\jlcue\Documents\mag-sistema\ref\Reporte 08 FEBRERO 2026 BASE VIDA Y GMM VF2 OK ULTIMO (1).xlsm"

def check_summary_details():
    wb = openpyxl.load_workbook(FILE_PATH, data_only=False)
    ws = wb["RESUMEN VIDA Y GMM"]
    # 2024 results: Col AR to AU (44 to 47)
    for r in [10, 14]: # Agent 644717 is 10, 47968 is 14
        row_letter = str(r)
        for col in range(43, 56):
            formula = ws.cell(row=r, column=col).value
            col_letter = openpyxl.utils.get_column_letter(col)
            print(f"Row {r} Col {col_letter}: Formula={formula}")
    wb.close()

check_summary_details()
