
import openpyxl
import os
import json

FILE_PATH = r"c:\Users\jlcue\Documents\mag-sistema\ref\Reporte 08 FEBRERO 2026 BASE VIDA Y GMM VF2 OK ULTIMO (1).xlsm"

def get_pivot_data():
    wb = openpyxl.load_workbook(FILE_PATH, read_only=True, data_only=True)
    ws = wb["RESUMEN VIDA Y GMM"]
    
    data = []
    # Row 10 to 150 (approx agents count)
    for r in range(10, 150):
        code = ws.cell(row=r, column=1).value
        if not code: continue
        
        row_dict = {
            "agente_code": str(code),
            # 2024: AR(44), AS(45), AT(46), AU(47), AV(48)
            # AS: New, AT: Sub, AR: Equiv, AV: Count
            "2024_new": float(ws.cell(row=r, column=45).value or 0),
            "2024_sub": float(ws.cell(row=r, column=46).value or 0),
            "2024_equiv": float(ws.cell(row=r, column=44).value or 0),
            "2024_count": int(ws.cell(row=r, column=48).value or 0),
            
            # 2025: AW(49), AX(50), AY(51), AZ(52), BA(53), BB(54), BC(55)
            # BA: New, BB: Sub, AY: Equiv, AW: Count
            "2025_new": float(ws.cell(row=r, column=53).value or 0),
            "2025_sub": float(ws.cell(row=r, column=54).value or 0),
            "2025_equiv": float(ws.cell(row=r, column=51).value or 0),
            "2025_count": int(ws.cell(row=r, column=49).value or 0),
        }
        data.append(row_dict)
    
    wb.close()
    return data

pivot_data = get_pivot_data()
with open("scripts/pivot_vida.json", "w") as f:
    json.dump(pivot_data, f, indent=2)
print(f"Extracted {len(pivot_data)} agents records to scripts/pivot_vida.json")
