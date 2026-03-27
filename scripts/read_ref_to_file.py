
import openpyxl
import os

FILE_PATH = r"c:\Users\jlcue\Documents\mag-sistema\ref\Reporte 08 FEBRERO 2026 BASE VIDA Y GMM VF2 OK ULTIMO (1).xlsm"
SHEET_NAME = "RESUMEN VIDA Y GMM"
OUT_FILE = r"c:\Users\jlcue\Documents\mag-sistema\tmp\ref_excel_data.txt"

if not os.path.exists(FILE_PATH):
    with open(OUT_FILE, "w") as f: f.write(f"ERROR: File not found: {FILE_PATH}")
else:
    wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
    ws = wb[SHEET_NAME]
    
    lines = []
    lines.append(f"--- DATA FROM {SHEET_NAME} ---")
    
    # 2024: AR (44) to AV (48)
    lines.append("\n--- DATA FOR 2024 (Row 8, AR-AV) ---")
    for col in range(44, 49):
        label = ws.cell(row=9, column=col).value
        val = ws.cell(row=8, column=col).value
        lines.append(f"Col {openpyxl.utils.get_column_letter(col)}: {label} => {val}")
        
    # 2025: AW (49) to BC (55)
    lines.append("\n--- DATA FOR 2025 (Row 8, AW-BC) ---")
    for col in range(49, 56):
        label = ws.cell(row=9, column=col).value
        val = ws.cell(row=8, column=col).value
        lines.append(f"Col {openpyxl.utils.get_column_letter(col)}: {label} => {val}")
    
    with open(OUT_FILE, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    
    wb.close()
    print("Done writing to " + OUT_FILE)
