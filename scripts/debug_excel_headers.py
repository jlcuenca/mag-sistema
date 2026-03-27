
import openpyxl
import os

FILE_PATH = r"c:\Users\jlcue\Documents\mag-sistema\ref\Reporte 08 FEBRERO 2026 BASE VIDA Y GMM VF2 OK ULTIMO (1).xlsm"
SHEET_NAME = "RESUMEN VIDA Y GMM"

if not os.path.exists(FILE_PATH):
    print("File not found")
else:
    wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
    ws = wb[SHEET_NAME]
    
    # 2024: AR to AV
    for col in range(40, 60): # Column 40 to 60 to be sure
        L9 = ws.cell(row=9, column=col).value
        L10 = ws.cell(row=10, column=col).value
        print(f"Col {openpyxl.utils.get_column_letter(col)}: R9='{L9}' R10='{L10}'")
    wb.close()
