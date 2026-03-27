
import openpyxl
import os

FILE_PATH = r"c:\Users\jlcue\Documents\mag-sistema\ref\Reporte 08 FEBRERO 2026 BASE VIDA Y GMM VF2 OK ULTIMO (1).xlsm"
wb = openpyxl.load_workbook(FILE_PATH, read_only=True)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    headers = [str(c.value).upper() for c in next(ws.iter_rows(max_row=1))]
    for h in headers:
        if 'PRIMA' in h or 'NETA' in h or 'MONTO' in h:
            print(f"FOUND IN {sheet_name}: {h}")
wb.close()
