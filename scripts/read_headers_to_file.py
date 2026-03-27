
import openpyxl
import os

FILE_PATH = r"c:\Users\jlcue\Documents\mag-sistema\ref\Reporte 08 FEBRERO 2026 BASE VIDA Y GMM VF2 OK ULTIMO (1).xlsm"
SHEET_NAME = "RESUMEN VIDA Y GMM"
OUT_FILE = r"c:\Users\jlcue\Documents\mag-sistema\scripts\headers_ref.txt"

if not os.path.exists(FILE_PATH):
    with open(OUT_FILE, "w") as f: f.write("File not found")
else:
    wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
    ws = wb[SHEET_NAME]
    
    with open(OUT_FILE, "w", encoding="utf-8") as f:
        # Col 40 to 65
        for col in range(40, 66):
            L9 = ws.cell(row=9, column=col).value
            L10 = ws.cell(row=10, column=col).value
            V8 = ws.cell(row=8, column=col).value
            col_letter = openpyxl.utils.get_column_letter(col)
            f.write(f"Col {col_letter}: R9='{L9}' R10='{L10}' R8='{V8}'\n")
    wb.close()
    print("Done")
