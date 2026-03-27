
import openpyxl
import os

FILE_PATH = r"c:\Users\jlcue\Documents\mag-sistema\ref\Reporte 08 FEBRERO 2026 BASE VIDA Y GMM VF2 OK ULTIMO (1).xlsm"

def get_agents():
    wb = openpyxl.load_workbook(FILE_PATH, read_only=True, data_only=True)
    ws = wb["RESUMEN VIDA Y GMM"]
    agents = []
    for r in range(10, 200):
        code = ws.cell(row=r, column=1).value
        if not code: break
        agents.append(str(code))
    wb.close()
    return agents

def final_match_2024():
    agents = get_agents()
    print(f"Checking for {len(agents)} agents...")
    
    wb = openpyxl.load_workbook(FILE_PATH, read_only=True, data_only=True)
    ws = wb["PRIMA PAGADA"]
    total = 0
    # 30: filter, 11: neta, 1: agente
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if i > 255000: break
        
        agente = str(row[0])
        filter_val = str(row[29])
        neta = float(row[10] or 0)
        
        if filter_val == "PRIMER AÑO 2024" and agente in agents:
            total += neta
            
    print(f"Total 2024 for MAG Agents: {total:,.2f}")
    wb.close()

final_match_2024()
