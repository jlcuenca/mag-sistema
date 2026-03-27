
import openpyxl
import os

FILE_PATH = r"c:\Users\jlcue\Documents\mag-sistema\ref\Reporte 08 FEBRERO 2026 BASE VIDA Y GMM VF2 OK ULTIMO (1).xlsm"

def check_summary_names():
    wb = openpyxl.load_workbook(FILE_PATH, read_only=True, data_only=True)
    ws = wb["RESUMEN VIDA Y GMM"]
    for r in range(8, 25):
        name = ws.cell(row=r, column=1).value
        # Values from Col 44 (AR) to 50
        vals = [ws.cell(row=r, column=c).value for c in range(48, 56)]
        print(f"Row {r}: Name='{name}' Vals={vals}")
    wb.close()

check_summary_names()
