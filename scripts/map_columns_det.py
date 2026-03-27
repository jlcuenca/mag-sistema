
import openpyxl
import os

FILE_PATH = r"c:\Users\jlcue\Documents\mag-sistema\ref\Reporte 08 FEBRERO 2026 BASE VIDA Y GMM VF2 OK ULTIMO (1).xlsm"
SHEET_NAME = "DET 2024"

wb = openpyxl.load_workbook(FILE_PATH, read_only=True)
ws = wb[SHEET_NAME]

headers = [c.value for c in ws[1]]
for i, h in enumerate(headers):
    print(f"{i+1}: {h}")
    
# Sample row 2
data = [c.value for c in ws[2]]
print("\nSample Data (Row 2):")
for i, val in enumerate(data):
    print(f"{i+1}: {val}")

wb.close()
