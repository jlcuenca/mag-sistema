
import openpyxl
import os

FILE_PATH = r"c:\Users\jlcue\Documents\mag-sistema\ref\Reporte 08 FEBRERO 2026 BASE VIDA Y GMM VF2 OK ULTIMO (1).xlsm"
SHEET_NAME = "RESUMEN VIDA Y GMM"

if not os.path.exists(FILE_PATH):
    print(f"ERROR: File not found: {FILE_PATH}")
else:
    wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
    ws = wb[SHEET_NAME]
    
    # 2024: AR (44) to AV (48)
    # 2025: AW (49) to BC (55)
    
    print("--- DATA FOR 2024 (Row 8, AR-AV) ---")
    for col in range(44, 49):
        label = ws.cell(row=9, column=col).value
        val = ws.cell(row=8, column=col).value
        print(f"2024 Col {openpyxl.utils.get_column_letter(col)}: {label} => {val}")
        
    print("\n--- DATA FOR 2025 (Row 8, AW-BC) ---")
    for col in range(49, 56):
        label = ws.cell(row=9, column=col).value
        val = ws.cell(row=8, column=col).value
        print(f"2025 Col {openpyxl.utils.get_column_letter(col)}: {label} => {val}")
    
    wb.close()
