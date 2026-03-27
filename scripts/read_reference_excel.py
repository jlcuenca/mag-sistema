
import openpyxl
import os

FILE_PATH = r"c:\Users\jlcue\Documents\mag-sistema\ref\Reporte 08 FEBRERO 2026 BASE VIDA Y GMM VF2 OK ULTIMO (1).xlsm"
SHEET_NAME = "RESUMEN VIDA Y GMM"

if not os.path.exists(FILE_PATH):
    print(f"File not found: {FILE_PATH}")
else:
    wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"Sheet not found: {SHEET_NAME}. Available: {wb.sheetnames}")
    else:
        ws = wb[SHEET_NAME]
        print(f"Reading {SHEET_NAME} for 2024 (AR-AV) and 2025 (AW-BC)")
        
        row_8 = []
        row_9 = []
        for col_idx in range(44, 56): # AR (44) to BC (55)? No, AR is 18 + 26 = 44th column.
            # A=1, R=18. 1*26 + 18 = 44. Correct.
            # AR (44), AS (45), AT(46), AU(47), AV(48) -> 2024
            # AW (49), AX (50), AY(51), AZ(52), BA(53), BB(54), BC(55) -> 2025
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            row_8.append(ws.cell(row=8, column=col_idx).value)
            row_9.append(ws.cell(row=9, column=col_idx).value)
        
        print("\nColumn Labels (Row 9):")
        print(row_9)
        print("\nValues (Row 8):")
        print(row_8)
    wb.close()
