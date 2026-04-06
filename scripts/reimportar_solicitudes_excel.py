"""
╔══════════════════════════════════════════════════════════════════╗
║  REIMPORTAR SOLICITUDES desde Excel Original                    ║
║  Fuente: DATA concentrado 26 OCT 2025 (27,491 filas, 66 cols)  ║
╚══════════════════════════════════════════════════════════════════╝

Lee la pestaña 'DATA concentrado 26 OCT 2025' del Excel original,
purga los datos existentes y reimporta con las 66 columnas completas.
Después aplica reglas S1-S7 en batch.

Uso: python scripts/reimportar_solicitudes_excel.py
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from datetime import datetime
from collections import Counter
from sqlalchemy import text
from api.database import SessionLocal, engine, Base
from api.rules_solicitudes import (
    normalizar_ramo, derivar_estado_de_etapa, detectar_solicitud_atorada,
    calcular_dias_tramite, evaluar_sla, calcular_tasa_conversion,
)

EXCEL_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "ref",
    "original base de VIDA Y GMM_DIC (V24 MARZO26)HF ACTUALIZADO3.xlsx"
)
SHEET_NAME = "DATA concentrado 26 OCT 2025"


def safe_str(val, max_len=None):
    """Convierte valor a string limpio o None"""
    if val is None:
        return None
    s = str(val).strip()
    if s in ('', 'None', 'nan', 'NaN', 'NaT', '-', '#N/A', '#', '#N/A!', '#REF!'):
        return None
    if max_len:
        s = s[:max_len]
    return s


def safe_int(val):
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def safe_float(val):
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def safe_date(val):
    """Convierte datetime/string a formato YYYY-MM-DD"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if s in ('', 'None', 'nan', '-', '#N/A'):
        return None
    return s[:10]


def banner(msg):
    print(f"\n{'═'*60}")
    print(f"  {msg}")
    print(f"{'═'*60}", flush=True)


def main():
    start = datetime.now()
    banner("REIMPORTACIÓN SOLICITUDES — Excel Original")
    print(f"  Archivo: {os.path.basename(EXCEL_PATH)}")
    print(f"  Pestaña: {SHEET_NAME}")
    print(f"  Inicio:  {start.strftime('%Y-%m-%d %H:%M:%S')}")

    if not os.path.exists(EXCEL_PATH):
        print(f"  ❌ No se encontró el archivo: {EXCEL_PATH}")
        return

    # ──────────────────────────────────────────────────────────
    # PASO 0: Asegurar columnas nuevas en la BD
    # ──────────────────────────────────────────────────────────
    banner("PASO 0: Verificando esquema de BD")
    db = SessionLocal()

    # Crear tablas si no existen
    Base.metadata.create_all(bind=engine)

    # Agregar columnas nuevas si no existen (SQLite ALTER TABLE)
    new_cols = {
        "agente_nombre": "VARCHAR(200)",
        "segmento": "VARCHAR(100)",
        "segmento_agrupado": "VARCHAR(50)",
        "gestion_comercial": "VARCHAR(200)",
        "poliza_6_digitos": "VARCHAR(20)",
        "primer_ano": "VARCHAR(50)",
        "dias_devengados": "INTEGER",
        "estatus_pago": "VARCHAR(50)",
        "sub_etapa": "VARCHAR(100)",
    }

    for col_name, col_type in new_cols.items():
        try:
            db.execute(text(f"ALTER TABLE solicitudes ADD COLUMN {col_name} {col_type}"))
            db.commit()
            print(f"  ✅ Columna '{col_name}' agregada")
        except Exception:
            db.rollback()
            # Ya existe, ok

    print("  ✅ Esquema verificado", flush=True)

    # ──────────────────────────────────────────────────────────
    # PASO 1: Leer Excel
    # ──────────────────────────────────────────────────────────
    banner("PASO 1: Leyendo Excel (esto puede tomar ~60s)")

    import openpyxl
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]

    # Leer headers
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, max_col=66):
        headers = [str(c.value).strip() if c.value else f'COL_{i}' for i, c in enumerate(row)]

    print(f"  Columnas leídas: {len(headers)}")

    # Leer todas las filas
    rows = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, max_col=66, values_only=True), 2):
        if row[0] is None:  # Skip empty rows (NOSOL is required)
            continue
        rows.append(row)

    wb.close()
    print(f"  ✅ {len(rows):,} filas leídas desde Excel", flush=True)

    # ──────────────────────────────────────────────────────────
    # PASO 2: Deduplicar por NOSOL (quedarnos con última etapa)
    # ──────────────────────────────────────────────────────────
    banner("PASO 2: Agrupando etapas por NOSOL")

    # Un NOSOL puede tener múltiples filas (una por etapa)
    # Para solicitudes: quedamos con la ÚLTIMA etapa
    # Para etapas: guardamos TODAS
    etapas_all = []     # Todas las filas como etapas
    sol_by_nosol = {}   # Última fila por nosol

    for row in rows:
        nosol = safe_str(row[0])
        if not nosol:
            continue

        # Guardar como etapa
        etapas_all.append(row)

        # Para solicitud, quedarnos con la más reciente por fecetapa
        fecetapa = safe_date(row[8])  # col I: FECETAPA
        if nosol not in sol_by_nosol:
            sol_by_nosol[nosol] = row
        else:
            existing_fecha = safe_date(sol_by_nosol[nosol][8])
            if fecetapa and (not existing_fecha or fecetapa > existing_fecha):
                sol_by_nosol[nosol] = row

    print(f"  Solicitudes únicas: {len(sol_by_nosol):,}")
    print(f"  Total etapas:       {len(etapas_all):,}", flush=True)

    # ──────────────────────────────────────────────────────────
    # PASO 3: Purgar tablas existentes
    # ──────────────────────────────────────────────────────────
    banner("PASO 3: Purgando datos existentes")

    count_sol = db.execute(text("SELECT COUNT(*) FROM solicitudes")).scalar()
    count_eta = db.execute(text("SELECT COUNT(*) FROM etapas_solicitudes")).scalar()

    db.execute(text("DELETE FROM etapas_solicitudes"))
    db.execute(text("DELETE FROM solicitudes"))
    db.commit()

    print(f"  Eliminadas: {count_sol:,} solicitudes, {count_eta:,} etapas", flush=True)

    # ──────────────────────────────────────────────────────────
    # PASO 4: Insertar SOLICITUDES (cabeceras únicas)
    # ──────────────────────────────────────────────────────────
    banner("PASO 4: Insertando solicitudes")

    now = datetime.now().isoformat()
    inserted = 0
    counters = Counter()

    for nosol, row in sol_by_nosol.items():
        # Mapeo Excel → DB
        # Cols: 0=NOSOL, 1=NOMRAMO, 2=FECRECEPCION, 3=CONTRATANTE
        # 4=DIA_RECEPCION, 5=MES_RECEPCION, 6=ANO_RECEPCION
        # 7=ETAPA, 8=FECETAPA, 9=DIA_ETAPA, 10=MES_ETAPA, 11=ANO_ETAPA
        # 12=OBSERVACIONES, 13=IDAGENTE, 14=NUEVO, 15=ANTAXA
        # 16=REINGRESO, 17=CONTASOL, 18=POLIZA, 19=NUMSOLICITANTES
        # 20=SYSDATE, 21=POLIZA 6 DIGITOS, 22=AGENTE, 23=PROMOTOR
        # 24=SEGMENTO, 25=SEGMENTO AGRUPADO, 26=GESTION COMERCIAL
        # 27=FECINI, 34=PRIMER AÑO VIDA, 54=PRIMER AÑO GMM
        # 55=DÍAS DEVENGADOS, 56=ESTATUS DE PAGO, 57=SUB ETAPA

        nomramo = safe_str(row[1])
        etapa = safe_str(row[7])
        poliza = safe_str(row[18])

        # S1: Normalizar ramo
        ramo_norm = normalizar_ramo(nomramo or '')
        # S2: Estado
        estado_result = derivar_estado_de_etapa(etapa or '', poliza or '')
        # S6: Días trámite
        fecrecepcion = safe_date(row[2])
        fecetapa = safe_date(row[8])
        dias = calcular_dias_tramite(fecrecepcion or '', fecetapa or '')
        # S4: Alerta
        alerta = detectar_solicitud_atorada(fecetapa, estado_result["estado"])
        # S7: SLA
        sla = evaluar_sla(dias)

        # Primer año: combinar Vida (col 34) y GMM (col 54)
        primer_ano = safe_str(row[34]) or safe_str(row[54])

        counters[estado_result["estado"]] += 1
        counters[ramo_norm or 'SIN_RAMO'] += 1

        db.execute(text("""
            INSERT INTO solicitudes (
                nosol, nomramo, fecrecepcion, contratante_nombre,
                dia_recepcion, mes_recepcion, ano_recepcion,
                idagente, nuevo, antaxa, reingreso, contasol,
                poliza_numero, numsolicitantes, fecha_sistema,
                agente_nombre, promotor_nombre, segmento, segmento_agrupado,
                gestion_comercial, poliza_6_digitos, primer_ano,
                dias_devengados, estatus_pago, sub_etapa,
                ultima_etapa, ultima_subetapa, fecha_ultima_etapa,
                observaciones_etapa,
                ramo_normalizado, estado, tipo_rechazo,
                dias_tramite, alerta_atorada, sla_cumplido,
                fuente, created_at, updated_at
            ) VALUES (
                :nosol, :nomramo, :fecrecepcion, :contratante,
                :dia_rec, :mes_rec, :ano_rec,
                :idagente, :nuevo, :antaxa, :reingreso, :contasol,
                :poliza, :numsolicitantes, :fecha_sistema,
                :agente_nombre, :promotor, :segmento, :segmento_agrupado,
                :gestion, :poliza_6, :primer_ano,
                :dias_devengados, :estatus_pago, :sub_etapa,
                :ultima_etapa, :ultima_subetapa, :fecha_ultima_etapa,
                :observaciones,
                :ramo_norm, :estado, :tipo_rechazo,
                :dias_tramite, :alerta, :sla,
                :fuente, :now, :now
            )
        """), {
            "nosol": nosol,
            "nomramo": nomramo,
            "fecrecepcion": fecrecepcion,
            "contratante": safe_str(row[3], 300),
            "dia_rec": safe_int(row[4]),
            "mes_rec": safe_int(row[5]),
            "ano_rec": safe_int(row[6]),
            "idagente": safe_str(row[13], 20),
            "nuevo": safe_int(row[14]),
            "antaxa": safe_int(row[15]),
            "reingreso": safe_int(row[16]),
            "contasol": safe_int(row[17]),
            "poliza": safe_str(row[18], 30),
            "numsolicitantes": safe_int(row[19]),
            "fecha_sistema": safe_date(row[20]),
            "agente_nombre": safe_str(row[22], 200),
            "promotor": safe_str(row[23], 200),
            "segmento": safe_str(row[24], 100),
            "segmento_agrupado": safe_str(row[25], 50),
            "gestion": safe_str(row[26], 200),
            "poliza_6": safe_str(row[21], 20),
            "primer_ano": safe_str(primer_ano, 50),
            "dias_devengados": safe_int(row[55]),
            "estatus_pago": safe_str(row[56], 50),
            "sub_etapa": safe_str(row[57], 100),
            "ultima_etapa": safe_str(row[7], 50),
            "ultima_subetapa": safe_str(row[57], 50),
            "fecha_ultima_etapa": fecetapa,
            "observaciones": safe_str(row[12]),
            "ramo_norm": ramo_norm,
            "estado": estado_result["estado"],
            "tipo_rechazo": estado_result.get("tipo_rechazo"),
            "dias_tramite": dias,
            "alerta": alerta,
            "sla": sla,
            "fuente": "EXCEL_ORIGINAL_V24",
            "now": now,
        })

        inserted += 1
        if inserted % 1000 == 0:
            db.commit()
            print(f"  ... {inserted:,} solicitudes insertadas", flush=True)

    db.commit()
    print(f"  ✅ {inserted:,} solicitudes insertadas total", flush=True)

    # ──────────────────────────────────────────────────────────
    # PASO 5: Insertar ETAPAS (todas las filas)
    # ──────────────────────────────────────────────────────────
    banner("PASO 5: Insertando etapas")

    # Primero construir el mapa nosol → solicitud_id
    nosol_id_map = {}
    for r in db.execute(text("SELECT nosol, id FROM solicitudes")).mappings():
        nosol_id_map[r["nosol"]] = r["id"]

    etapa_count = 0
    for row in etapas_all:
        nosol = safe_str(row[0])
        if not nosol:
            continue

        sol_id = nosol_id_map.get(nosol)
        fecrecepcion = safe_date(row[2])
        fecetapa = safe_date(row[8])
        dias = calcular_dias_tramite(fecrecepcion or '', fecetapa or '')

        db.execute(text("""
            INSERT INTO etapas_solicitudes (
                nosol, solicitud_id, nomramo, fecrecepcion, contratante,
                dia_recepcion, mes_recepcion, ano_recepcion,
                etapa, subetapa, fecetapa, dia_etapa, mes_etapa, ano_etapa,
                observaciones, idagente, nuevo, antaxa, reingreso,
                contasol, poliza, numsolicitantes, fecha_sistema,
                dias_tramite, fuente, created_at
            ) VALUES (
                :nosol, :sol_id, :nomramo, :fecrecepcion, :contratante,
                :dia_rec, :mes_rec, :ano_rec,
                :etapa, :subetapa, :fecetapa, :dia_etapa, :mes_etapa, :ano_etapa,
                :observaciones, :idagente, :nuevo, :antaxa, :reingreso,
                :contasol, :poliza, :numsolicitantes, :fecha_sistema,
                :dias_tramite, :fuente, :now
            )
        """), {
            "nosol": nosol,
            "sol_id": sol_id,
            "nomramo": safe_str(row[1]),
            "fecrecepcion": fecrecepcion,
            "contratante": safe_str(row[3], 300),
            "dia_rec": safe_int(row[4]),
            "mes_rec": safe_int(row[5]),
            "ano_rec": safe_int(row[6]),
            "etapa": safe_str(row[7], 50),
            "subetapa": safe_str(row[57], 50),
            "fecetapa": fecetapa,
            "dia_etapa": safe_int(row[9]),
            "mes_etapa": safe_int(row[10]),
            "ano_etapa": safe_int(row[11]),
            "observaciones": safe_str(row[12]),
            "idagente": safe_str(row[13], 20),
            "nuevo": safe_int(row[14]),
            "antaxa": safe_int(row[15]),
            "reingreso": safe_int(row[16]),
            "contasol": safe_int(row[17]),
            "poliza": safe_str(row[18], 30),
            "numsolicitantes": safe_int(row[19]),
            "fecha_sistema": safe_date(row[20]),
            "dias_tramite": dias,
            "fuente": "EXCEL_ORIGINAL_V24",
            "now": now,
        })

        etapa_count += 1
        if etapa_count % 5000 == 0:
            db.commit()
            print(f"  ... {etapa_count:,} etapas insertadas", flush=True)

    db.commit()
    print(f"  ✅ {etapa_count:,} etapas insertadas total", flush=True)

    # ──────────────────────────────────────────────────────────
    # PASO 6: Tasa de conversión por agente (S5)
    # ──────────────────────────────────────────────────────────
    banner("PASO 6: Calculando tasa de conversión por agente")

    agentes_stats = db.execute(text("""
        SELECT idagente,
               COUNT(*) as total,
               SUM(CASE WHEN estado IN ('EMITIDA','PAGADA') THEN 1 ELSE 0 END) as emitidas
        FROM solicitudes
        WHERE idagente IS NOT NULL AND idagente != ''
        GROUP BY idagente
    """)).mappings().all()

    for a in agentes_stats:
        tasa = calcular_tasa_conversion(a["total"], a["emitidas"])
        db.execute(text(
            "UPDATE solicitudes SET tasa_conversion_agente = :tasa WHERE idagente = :id"
        ), {"tasa": tasa, "id": a["idagente"]})
    db.commit()
    print(f"  ✅ {len(agentes_stats):,} agentes con tasa calculada", flush=True)

    # ──────────────────────────────────────────────────────────
    # RESUMEN
    # ──────────────────────────────────────────────────────────
    elapsed = (datetime.now() - start).total_seconds()
    total_sol = db.execute(text("SELECT COUNT(*) FROM solicitudes")).scalar()
    total_eta = db.execute(text("SELECT COUNT(*) FROM etapas_solicitudes")).scalar()

    # Por estado
    estados = db.execute(text("""
        SELECT estado, COUNT(*) as cnt FROM solicitudes GROUP BY estado ORDER BY cnt DESC
    """)).mappings().all()

    # Por ramo
    ramos = db.execute(text("""
        SELECT ramo_normalizado, COUNT(*) as cnt FROM solicitudes GROUP BY ramo_normalizado ORDER BY cnt DESC
    """)).mappings().all()

    # Por año
    anos = db.execute(text("""
        SELECT ano_recepcion, COUNT(*) as cnt FROM solicitudes
        WHERE ano_recepcion IS NOT NULL GROUP BY ano_recepcion ORDER BY ano_recepcion
    """)).mappings().all()

    banner("RESUMEN FINAL")
    print(f"""
  📋 Solicitudes:  {total_sol:,}
  ⏱️ Etapas:       {total_eta:,}

  ── Por Estado ──────────────────""")
    for e in estados:
        print(f"    {e['estado'] or 'NULL':15s} → {e['cnt']:,}")
    print(f"\n  ── Por Ramo ────────────────────")
    for r in ramos:
        print(f"    {r['ramo_normalizado'] or 'NULL':10s} → {r['cnt']:,}")
    print(f"\n  ── Por Año ─────────────────────")
    for a in anos:
        print(f"    {a['ano_recepcion'] or '?':10} → {a['cnt']:,}")
    print(f"\n  ⏱️ Tiempo total: {elapsed:.1f}s")

    db.close()


if __name__ == "__main__":
    main()
