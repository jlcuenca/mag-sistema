
import openpyxl
import os

FILE_PATH = r"c:\Users\jlcue\Documents\mag-sistema\ref\Reporte 08 FEBRERO 2026 BASE VIDA Y GMM VF2 OK ULTIMO (1).xlsm"
SHEET_NAME = "DET 2024"

wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
ws = wb[SHEET_NAME]

print("--- ROW 1 (HEADERS) ---")
print([ws.cell(row=1, column=c).value for c in range(1, 42)])

print("\n--- ROW 2 (DATA) ---")
print([ws.cell(row=2, column=c).value for c in range(1, 42)])

wb.close()
