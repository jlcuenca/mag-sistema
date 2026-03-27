
import openpyxl
import os

FILE_PATH = r"c:\Users\jlcue\Documents\mag-sistema\ref\Reporte 08 FEBRERO 2026 BASE VIDA Y GMM VF2 OK ULTIMO (1).xlsm"

def check_global_filters():
    wb = openpyxl.load_workbook(FILE_PATH, read_only=True, data_only=True)
    ws = wb["RESUMEN VIDA Y GMM"]
    print(f"B1: {ws['B1'].value}")
    print(f"B2: {ws['B2'].value}")
    print(f"B3: {ws['B3'].value}")
    wb.close()

check_global_filters()
