
import openpyxl
import os

FILE_PATH = r"c:\Users\jlcue\Documents\mag-sistema\ref\Reporte 08 FEBRERO 2026 BASE VIDA Y GMM VF2 OK ULTIMO (1).xlsm"

def check_summary_formulas():
    wb = openpyxl.load_workbook(FILE_PATH, data_only=False) # Get formulas
    ws = wb["RESUMEN VIDA Y GMM"]
    
    # 2024 results: Col AR to AV (44 to 48)
    for col in range(44, 56):
        formula = ws.cell(row=8, column=col).value
        col_letter = openpyxl.utils.get_column_letter(col)
        print(f"Col {col_letter}: Formula={formula}")
    wb.close()

check_summary_formulas()
