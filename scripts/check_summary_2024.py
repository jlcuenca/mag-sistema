
import openpyxl
import os

FILE_PATH = r"c:\Users\jlcue\Documents\mag-sistema\ref\Reporte 08 FEBRERO 2026 BASE VIDA Y GMM VF2 OK ULTIMO (1).xlsm"

def check_summary_2024():
    wb = openpyxl.load_workbook(FILE_PATH, read_only=True, data_only=True)
    ws = wb["RESUMEN VIDA Y GMM"]
    print("AGENT CODE | POL NUEVAS | EQUIV | PRIMA NUEVA | PRIMA SUB | TOTAL (2024)")
    for r in range(10, 25):
        key = ws.cell(row=r, column=1).value
        # 2024: AR(44) to AV(48)
        # AR: EQUIV, AS: PRIMA NEW, AT: PRIMA SUB, AU: TOTAL, AV: N POL NUEVAS
        v_equiv = ws.cell(row=r, column= openpyxl.utils.column_index_from_string('AR')).value
        v_new   = ws.cell(row=r, column= openpyxl.utils.column_index_from_string('AS')).value
        v_sub   = ws.cell(row=r, column= openpyxl.utils.column_index_from_string('AT')).value
        v_tot   = ws.cell(row=r, column= openpyxl.utils.column_index_from_string('AU')).value
        v_count = ws.cell(row=r, column= openpyxl.utils.column_index_from_string('AV')).value
        print(f"{key} | {v_count} | {v_equiv} | {v_new:,.2f} | {v_sub:,.2f} | {v_tot:,.2f}")
    wb.close()

check_summary_2024()
