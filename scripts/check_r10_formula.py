
import openpyxl
import os

FILE_PATH = r"c:\Users\jlcue\Documents\mag-sistema\ref\Reporte 08 FEBRERO 2026 BASE VIDA Y GMM VF2 OK ULTIMO (1).xlsm"

def check_r10_formula():
    wb = openpyxl.load_workbook(FILE_PATH, read_only=True, data_only=False)
    ws = wb["RESUMEN VIDA Y GMM"]
    # Check Agente 644717 (Row 10)
    # AS: Prima New, AT: Prima Sub
    print(f"AS10: {ws['AS10'].value}")
    print(f"AT10: {ws['AT10'].value}")
    wb.close()

check_r10_formula()
